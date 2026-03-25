import json
import os
import re
import shutil
import subprocess
import sys
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
EXCEL_ERROR_TOKENS = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!",
    "#N/A", "#SPILL!", "#CALC!", "#FIELD!", "#UNKNOWN!", "#GETTING_DATA",
}
EXCEL_SHEET_INVALID_CHARS = re.compile(r"[:\\/?*\[\]]")

_PHONE_RE = re.compile(r"^(0\d{1,2})[-.\s]?(\d{3,4})[-.\s]?(\d{4})$")
_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
_DATE_RE = re.compile(r"^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$")

SIMPLE_RULE_TYPES = ["필수 입력", "숫자만", "전화번호", "이메일", "날짜"]

APP_FONT_FAMILY = "Malgun Gothic" if os.name == "nt" else "DejaVu Sans"
THEME_POLL_INTERVAL_MS = 2000
ICON_FILE_BASENAME = "Excel_Merge_ICON"
THEME_PALETTES = {
    "light": {
        "bg": "#f3f5f8",
        "surface": "#ffffff",
        "surface_alt": "#e8edf6",
        "panel": "#f7f9fc",
        "border": "#d6dde8",
        "text": "#141821",
        "muted": "#5d6678",
        "accent": "#2c64f2",
        "accent_hover": "#1f52d5",
        "accent_text": "#ffffff",
        "selection": "#d9e7ff",
        "selection_text": "#0f1728",
        "input_bg": "#fbfcff",
        "tree_heading": "#edf2fa",
        "success": "#1f8f5f",
        "warning": "#b86a00",
        "error": "#c23b33",
    },
    "dark": {
        "bg": "#101216",
        "surface": "#171a20",
        "surface_alt": "#1e232c",
        "panel": "#11151b",
        "border": "#2b313d",
        "text": "#e8edf5",
        "muted": "#9ca7b8",
        "accent": "#63a0ff",
        "accent_hover": "#87b5ff",
        "accent_text": "#081120",
        "selection": "#2a3e5d",
        "selection_text": "#f4f8ff",
        "input_bg": "#0f1318",
        "tree_heading": "#20252e",
        "success": "#67d7a2",
        "warning": "#ffbe63",
        "error": "#ff7b72",
    },
}


def _read_windows_apps_use_light_theme() -> Optional[int]:
    reg_path = r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
    value_name = "AppsUseLightTheme"

    if sys.platform == "win32":
        try:
            import winreg  # type: ignore

            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path) as key:
                value, _ = winreg.QueryValueEx(key, value_name)
                return int(value)
        except Exception:
            return None

    reg_exe = shutil.which("reg.exe")
    if not reg_exe:
        fallback = "/mnt/c/Windows/System32/reg.exe"
        if os.path.exists(fallback):
            reg_exe = fallback
    if not reg_exe:
        return None

    try:
        result = subprocess.run(
            [reg_exe, "query", f"HKCU\\{reg_path}", "/v", value_name],
            capture_output=True,
            text=True,
            timeout=2,
            check=False,
        )
    except Exception:
        return None

    if result.returncode != 0:
        return None

    match = re.search(rf"{value_name}\s+REG_DWORD\s+0x([0-9a-fA-F]+)", result.stdout)
    if not match:
        return None
    return int(match.group(1), 16)


def detect_system_theme_mode() -> str:
    value = _read_windows_apps_use_light_theme()
    if value is None:
        return "light"
    return "light" if value else "dark"


def get_palette(theme_mode: str) -> dict:
    return THEME_PALETTES.get(theme_mode, THEME_PALETTES["light"])


def apply_window_palette(window: tk.Misc, palette: dict):
    try:
        window.configure(bg=palette["bg"])
    except tk.TclError:
        pass


def get_app_base_dir() -> str:
    if getattr(sys, "frozen", False):
        return getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(*parts: str) -> str:
    return os.path.join(get_app_base_dir(), *parts)


@dataclass
class ValidationRule:
    column: str
    rule_type: str
    value1: str = ""
    value2: str = ""
    value3: str = ""
    error_msg: str = ""  # 사용자 지정 오류 메시지 (비우면 자동 생성)

    def display(self) -> str:
        if self.rule_type == "숫자 범위":
            parts = []
            if self.value1:
                parts.append(f"{self.value1}이상")
            if self.value2:
                parts.append(f"{self.value2}이하")
            return f"숫자 범위({', '.join(parts)})" if parts else "숫자 범위"
        elif self.rule_type == "텍스트 포함":
            return f"'{self.value1}' 포함"
        elif self.rule_type == "허용값 목록":
            return f"허용값({self.value1})"
        elif self.rule_type == "조건부 검증":
            return f"[{self.value1}]={self.value2} → 허용값({self.value3})"
        return self.rule_type

    def to_dict(self) -> dict:
        return {"column": self.column, "rule_type": self.rule_type,
                "value1": self.value1, "value2": self.value2,
                "value3": self.value3, "error_msg": self.error_msg}

    @staticmethod
    def from_dict(d: dict) -> "ValidationRule":
        return ValidationRule(
            column=d["column"], rule_type=d["rule_type"],
            value1=d.get("value1", ""), value2=d.get("value2", ""),
            value3=d.get("value3", ""), error_msg=d.get("error_msg", ""),
        )


def is_empty_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def is_excel_error_value(value) -> bool:
    return isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_TOKENS


def normalize_headers(header_values: List[object]) -> List[str]:
    normalized = []
    used = {}
    for idx, value in enumerate(header_values, start=1):
        name = str(value).strip() if value is not None else ""
        if not name:
            name = f"Column_{idx}"
        if name in used:
            used[name] += 1
            name = f"{name}_{used[name]}"
        else:
            used[name] = 1
        normalized.append(name)
    return normalized


def make_excel_sheet_name(name: str) -> str:
    candidate = EXCEL_SHEET_INVALID_CHARS.sub("_", (name or "").strip())
    candidate = candidate.strip("'")
    candidate = candidate or "sheet"
    return candidate[:31]


def unique_sheet_name(base_name: str, existing_names: set) -> str:
    candidate = make_excel_sheet_name(base_name)
    if candidate not in existing_names:
        return candidate
    counter = 2
    while True:
        suffix = f"_{counter}"
        max_base_len = 31 - len(suffix)
        candidate = f"{make_excel_sheet_name(base_name)[:max_base_len]}{suffix}"
        if candidate not in existing_names:
            return candidate
        counter += 1


def read_excel_data(
    file_path: str, sheet_name: Optional[str],
    header_row: int, skip_empty_rows: bool,
) -> Tuple[pd.DataFrame, str]:
    wb = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"시트 '{sheet_name}'을(를) 찾을 수 없습니다.")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False):
            if not row:
                continue
            rows.append((row[0].row, [cell.value for cell in row]))
        if not rows:
            raise ValueError("데이터가 없습니다.")
        if header_row < 1 or header_row > len(rows):
            raise ValueError(f"헤더 행 번호가 유효하지 않습니다. (1~{len(rows)})")
        headers = normalize_headers(rows[header_row - 1][1])
        records = []
        for row_index, values in rows[header_row:]:
            if skip_empty_rows and all(is_empty_value(v) for v in values):
                continue
            row_dict = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            row_dict["__source_row__"] = row_index
            records.append(row_dict)
        return pd.DataFrame(records, columns=headers + ["__source_row__"]), ws.title
    finally:
        wb.close()


def _check_rule(rule, value):
    """규칙 위반 시 기본 메시지를 반환, 위반 아니면 None"""
    empty = is_empty_value(value)
    if rule.rule_type == "필수 입력":
        if empty:
            return "필수값이 비어 있습니다."
        return None
    if empty:
        return None
    tv = str(value).strip()
    if rule.rule_type == "숫자만":
        try:
            float(tv.replace(",", ""))
        except ValueError:
            return f"숫자가 아닙니다: {tv}"
    elif rule.rule_type == "숫자 범위":
        try:
            num = float(tv.replace(",", ""))
        except ValueError:
            return f"숫자가 아닙니다: {tv}"
        mn = float(rule.value1) if rule.value1 else None
        mx = float(rule.value2) if rule.value2 else None
        if mn is not None and num < mn:
            return f"{mn} 미만: {tv}"
        if mx is not None and num > mx:
            return f"{mx} 초과: {tv}"
    elif rule.rule_type == "텍스트 포함":
        if rule.value1 and rule.value1 not in tv:
            return f"'{rule.value1}' 미포함: {tv}"
    elif rule.rule_type == "허용값 목록":
        allowed = {v.strip() for v in rule.value1.split(",") if v.strip()}
        if allowed and tv not in allowed:
            return f"허용값에 없음: {tv}"
    elif rule.rule_type == "전화번호":
        if not _PHONE_RE.match(tv.replace(" ", "")):
            return f"전화번호 형식 아님: {tv}"
    elif rule.rule_type == "이메일":
        if not _EMAIL_RE.match(tv):
            return f"이메일 형식 아님: {tv}"
    elif rule.rule_type == "날짜":
        if isinstance(value, datetime):
            return None
        m = _DATE_RE.match(tv)
        if not m:
            return f"날짜 형식 아님: {tv}"
        elif not (1 <= int(m.group(2)) <= 12 and 1 <= int(m.group(3)) <= 31):
            return f"유효하지 않은 날짜: {tv}"
    return None


def validate_rules_by_row(df, rules):
    """행 번호별 오류 메시지 dict 반환: {source_row: ["메시지1", ...]}"""
    row_messages: Dict[int, List[str]] = {}
    if not rules:
        return row_messages

    # 빈 칸 / 오류값도 포함
    data_cols = [c for c in df.columns if c != "__source_row__"]
    for _, row in df.iterrows():
        src = row.get("__source_row__")
        if src is None:
            continue
        messages = []

        # 빈 칸 / 오류값 자동 검출
        for col in data_cols:
            v = row[col]
            if is_excel_error_value(v):
                messages.append(f"[{col}] 오류값: {v}")

        # 규칙 검증
        for column, rule in rules.items():
            if column not in df.columns:
                continue
            if rule.rule_type == "조건부 검증":
                # value1=조건열, value2=조건값, value3=허용값(쉼표 구분)
                cond_col = rule.value1
                if cond_col not in df.columns:
                    continue
                cond_val = str(row[cond_col]).strip() if not is_empty_value(row[cond_col]) else ""
                if cond_val != rule.value2.strip():
                    continue  # 조건 불일치 → 이 규칙 적용 안 함
                target_val = str(row[column]).strip() if not is_empty_value(row[column]) else ""
                allowed = {v.strip() for v in rule.value3.split(",") if v.strip()}
                if target_val not in allowed:
                    default_msg = f"[{cond_col}]={cond_val}일 때 허용값({rule.value3})이어야 함: {target_val}"
                    msg = rule.error_msg if rule.error_msg else default_msg
                    messages.append(f"[{column}] {msg}")
                continue
            default_msg = _check_rule(rule, row[column])
            if default_msg:
                msg = rule.error_msg if rule.error_msg else default_msg
                messages.append(f"[{column}] {msg}")

        if messages:
            row_messages[int(src)] = messages

    return row_messages


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  GUI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ExcelMergeApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("엑셀 시트 병합 도구")
        self.root.geometry("1230x860")
        self.root.minsize(1000, 700)
        self.style = ttk.Style(self.root)
        self.theme_mode = "light"
        self.theme_job = None
        self.palette = get_palette(self.theme_mode)
        self._app_icon_image = None

        self.selected_files: List[str] = []
        # 파일별 규칙: {파일경로: {열이름: ValidationRule}}
        self.file_rules: Dict[str, Dict[str, ValidationRule]] = {}
        # 파일별 선택된 시트: {파일경로: [시트명, ...]}  (비어있으면 전체)
        self.file_sheets: Dict[str, List[str]] = {}
        self.file_available_sheets: Dict[str, List[str]] = {}
        self.preview_columns: List[str] = []
        self.preview_display_columns: List[str] = []
        self.current_preview_file: Optional[str] = None

        self.target_path_var = tk.StringVar()
        self.header_row_var = tk.IntVar(value=1)
        self.skip_empty_rows_var = tk.BooleanVar(value=True)
        self.theme_status_var = tk.StringVar()
        self.preview_status_var = tk.StringVar(value="선택된 파일 없음")
        self._sheet_check_widgets: list = []

        self._apply_app_icon()
        self._build_ui()
        self._refresh_bulk_sheet_controls()
        self._sync_theme(force=True)
        self._schedule_theme_sync()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        if self.theme_job:
            self.root.after_cancel(self.theme_job)
            self.theme_job = None
        self.root.destroy()

    def _schedule_theme_sync(self):
        self.theme_job = self.root.after(THEME_POLL_INTERVAL_MS, self._poll_theme)

    def _poll_theme(self):
        self._sync_theme()
        self._schedule_theme_sync()

    def _sync_theme(self, force: bool = False):
        detected_mode = detect_system_theme_mode()
        if force or detected_mode != self.theme_mode:
            self.theme_mode = detected_mode
            self.palette = get_palette(self.theme_mode)
            self.root._excel_palette = self.palette
            self._apply_theme()

    def _apply_app_icon(self):
        ico_path = resource_path(f"{ICON_FILE_BASENAME}.ico")
        png_path = resource_path(f"{ICON_FILE_BASENAME}.png")
        svg_path = resource_path(f"{ICON_FILE_BASENAME}.svg")

        if os.name == "nt" and os.path.exists(ico_path):
            try:
                self.root.iconbitmap(default=ico_path)
            except tk.TclError:
                pass

        for candidate in (png_path, svg_path):
            if not os.path.exists(candidate):
                continue
            try:
                self._app_icon_image = tk.PhotoImage(file=candidate)
                self.root.iconphoto(True, self._app_icon_image)
                break
            except tk.TclError:
                continue

    def _apply_theme(self):
        palette = self.palette
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        base_font = (APP_FONT_FAMILY, 10)
        bold_font = (APP_FONT_FAMILY, 10, "bold")
        title_font = (APP_FONT_FAMILY, 18, "bold")
        button_font = (APP_FONT_FAMILY, 11, "bold")

        apply_window_palette(self.root, palette)
        self.root.option_add("*Font", "TkDefaultFont")

        self.style.configure(".", font=base_font)
        self.style.configure("TFrame", background=palette["bg"])
        self.style.configure("App.TFrame", background=palette["bg"])
        self.style.configure("Hero.TFrame", background=palette["surface_alt"])
        self.style.configure("Card.TFrame", background=palette["surface"])
        self.style.configure("Inset.TFrame", background=palette["panel"])

        self.style.configure("TLabel", background=palette["bg"], foreground=palette["text"], font=base_font)
        self.style.configure("Card.TLabel", background=palette["surface"], foreground=palette["text"], font=base_font)
        self.style.configure("CardHint.TLabel", background=palette["surface"], foreground=palette["muted"], font=base_font)
        self.style.configure("SectionTitle.TLabel", background=palette["surface"], foreground=palette["text"], font=bold_font)
        self.style.configure("Title.TLabel", background=palette["surface_alt"], foreground=palette["text"], font=title_font)
        self.style.configure("Subtitle.TLabel", background=palette["surface_alt"], foreground=palette["muted"], font=base_font)
        self.style.configure(
            "Badge.TLabel",
            background=palette["panel"],
            foreground=palette["accent"],
            font=bold_font,
            padding=(10, 6),
        )
        self.style.configure("Hint.TLabel", background=palette["bg"], foreground=palette["muted"], font=base_font)
        self.style.configure("AccentText.TLabel", background=palette["bg"], foreground=palette["accent"], font=base_font)

        self.style.configure(
            "Card.TLabelframe",
            background=palette["surface"],
            bordercolor=palette["border"],
            relief="solid",
            borderwidth=1,
        )
        self.style.configure(
            "Card.TLabelframe.Label",
            background=palette["surface"],
            foreground=palette["text"],
            font=bold_font,
        )

        self.style.configure(
            "TEntry",
            fieldbackground=palette["input_bg"],
            background=palette["input_bg"],
            foreground=palette["text"],
            insertcolor=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=8,
        )
        self.style.configure(
            "TSpinbox",
            fieldbackground=palette["input_bg"],
            background=palette["input_bg"],
            foreground=palette["text"],
            arrowcolor=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=4,
        )
        self.style.configure(
            "TCombobox",
            fieldbackground=palette["input_bg"],
            background=palette["input_bg"],
            foreground=palette["text"],
            arrowcolor=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=6,
        )
        self.style.map(
            "TCombobox",
            fieldbackground=[("readonly", palette["input_bg"])],
            foreground=[("readonly", palette["text"])],
        )

        self.style.configure(
            "TButton",
            background=palette["surface_alt"],
            foreground=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=(12, 8),
        )
        self.style.map(
            "TButton",
            background=[("active", palette["panel"]), ("pressed", palette["selection"])],
            foreground=[("disabled", palette["muted"])],
        )
        self.style.configure(
            "Accent.TButton",
            background=palette["accent"],
            foreground=palette["accent_text"],
            bordercolor=palette["accent"],
            lightcolor=palette["accent"],
            darkcolor=palette["accent"],
            font=button_font,
            padding=(14, 11),
        )
        self.style.map(
            "Accent.TButton",
            background=[("active", palette["accent_hover"]), ("pressed", palette["accent_hover"])],
            foreground=[("disabled", palette["muted"])],
        )

        self.style.configure("TCheckbutton", background=palette["surface"], foreground=palette["text"])
        self.style.configure("Panel.TCheckbutton", background=palette["panel"], foreground=palette["text"])
        self.style.map(
            "TCheckbutton",
            background=[("active", palette["surface"])],
            foreground=[("disabled", palette["muted"])],
        )
        self.style.map(
            "Panel.TCheckbutton",
            background=[("active", palette["panel"])],
            foreground=[("disabled", palette["muted"])],
        )
        self.style.configure(
            "TScrollbar",
            background=palette["surface_alt"],
            troughcolor=palette["panel"],
            bordercolor=palette["border"],
            arrowcolor=palette["text"],
        )

        self.style.configure(
            "Data.Treeview",
            background=palette["input_bg"],
            fieldbackground=palette["input_bg"],
            foreground=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            rowheight=28,
        )
        self.style.map(
            "Data.Treeview",
            background=[("selected", palette["selection"])],
            foreground=[("selected", palette["selection_text"])],
        )
        self.style.configure(
            "Data.Treeview.Heading",
            background=palette["tree_heading"],
            foreground=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            font=bold_font,
            padding=(8, 6),
        )
        self.style.map(
            "Data.Treeview.Heading",
            background=[("active", palette["surface_alt"])],
        )

        self.theme_status_var.set(
            f"Windows 테마 동기화: {'라이트 모드' if self.theme_mode == 'light' else '다크 모드'}"
        )

        self.style.configure(
            "Files.Treeview",
            background=palette["input_bg"],
            fieldbackground=palette["input_bg"],
            foreground=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            rowheight=30,
        )
        self.style.map(
            "Files.Treeview",
            background=[("selected", palette["selection"])],
            foreground=[("selected", palette["selection_text"])],
        )
        self.style.configure(
            "Files.Treeview.Heading",
            background=palette["tree_heading"],
            foreground=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            font=bold_font,
            padding=(8, 6),
        )
        self.style.map(
            "Files.Treeview.Heading",
            background=[("active", palette["surface_alt"])],
        )
        self.log_text.configure(
            bg=palette["input_bg"],
            fg=palette["text"],
            insertbackground=palette["text"],
            selectbackground=palette["selection"],
            selectforeground=palette["selection_text"],
            highlightbackground=palette["border"],
            highlightthickness=1,
            relief="flat",
            borderwidth=0,
        )
        if hasattr(self.log_text, "vbar"):
            self.log_text.vbar.configure(
                bg=palette["surface_alt"],
                troughcolor=palette["panel"],
                activebackground=palette["selection"],
                highlightbackground=palette["border"],
                relief="flat",
                borderwidth=0,
            )
        self.log_text.tag_configure("error", foreground=palette["error"])
        self.log_text.tag_configure("warn", foreground=palette["warning"])
        self.log_text.tag_configure("ok", foreground=palette["success"])
        self.log_text.tag_configure("info", foreground=palette["text"])
        self.ctx_menu.configure(
            bg=palette["surface"],
            fg=palette["text"],
            activebackground=palette["selection"],
            activeforeground=palette["selection_text"],
            relief="flat",
            borderwidth=1,
        )
    def _build_ui(self):
        main = ttk.Frame(self.root, style="App.TFrame", padding=(14, 14, 14, 14))
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=0, minsize=360)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(1, weight=1)

        header = ttk.Frame(main, style="Hero.TFrame", padding=(18, 14, 18, 14))
        header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="엑셀 시트 병합 도구", style="Title.TLabel").grid(
            row=0, column=0, sticky="w"
        )

        left = ttk.Frame(main, style="App.TFrame")
        left.grid(row=1, column=0, sticky="nsew", padx=(0, 12))
        left.columnconfigure(0, weight=1)

        right = ttk.Frame(main, style="App.TFrame")
        right.grid(row=1, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=5)
        right.rowconfigure(1, weight=3)

        s1 = ttk.LabelFrame(left, text="1. 결과 저장 파일", style="Card.TLabelframe", padding=(12, 10))
        s1.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        s1.columnconfigure(0, weight=1)
        ttk.Label(
            s1,
            text="병합 결과를 저장할 파일입니다.",
            style="CardHint.TLabel",
            wraplength=300,
            justify="left",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(s1, textvariable=self.target_path_var).grid(row=1, column=0, sticky="ew", pady=(0, 8))
        target_buttons = ttk.Frame(s1, style="Card.TFrame")
        target_buttons.grid(row=2, column=0, sticky="ew")
        target_buttons.columnconfigure(0, weight=1)
        target_buttons.columnconfigure(1, weight=1)
        ttk.Button(target_buttons, text="기존 파일 열기", command=self.select_target).grid(
            row=0, column=0, sticky="ew", padx=(0, 6)
        )
        ttk.Button(target_buttons, text="새 파일 만들기", command=self.create_new_target).grid(
            row=0, column=1, sticky="ew"
        )

        s2 = ttk.LabelFrame(left, text="2. 병합할 원본 파일", style="Card.TLabelframe", padding=(12, 10))
        s2.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        s2.columnconfigure(0, weight=1)
        s2.rowconfigure(1, weight=0, minsize=108)
        ttk.Label(
            s2,
            text="병합할 원본 엑셀 파일을 추가합니다.",
            style="CardHint.TLabel",
            wraplength=300,
            justify="left",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        list_frame = ttk.Frame(s2, style="Inset.TFrame", padding=(2, 2, 2, 2))
        list_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        list_frame.configure(height=108)
        list_frame.grid_propagate(False)
        self.file_listbox = ttk.Treeview(
            list_frame,
            style="Files.Treeview",
            columns=("file",),
            show="headings",
            selectmode="extended",
            height=3,
        )
        self.file_listbox.grid(row=0, column=0, sticky="nsew")
        self.file_listbox.heading("file", text="추가된 파일 목록")
        self.file_listbox.column("file", anchor="w", stretch=True, width=320)
        file_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_listbox.yview)
        file_scroll.grid(row=0, column=1, sticky="ns")
        self.file_listbox.config(yscrollcommand=file_scroll.set)
        self.file_listbox.bind("<<TreeviewSelect>>", self._on_file_select)

        btns = ttk.Frame(s2, style="Card.TFrame")
        btns.grid(row=2, column=0, sticky="ew", pady=(8, 8))
        btns.columnconfigure(0, weight=1)
        btns.columnconfigure(1, weight=1)
        btns.columnconfigure(2, weight=1)
        ttk.Button(btns, text="파일 추가", command=self.add_files).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(btns, text="선택 제거", command=self.remove_selected_files).grid(
            row=0, column=1, sticky="ew", padx=(0, 6)
        )
        ttk.Button(btns, text="전체 제거", command=self.clear_files).grid(row=0, column=2, sticky="ew")

        sheet_frame = ttk.Frame(s2, style="Card.TFrame")
        sheet_frame.grid(row=3, column=0, sticky="ew")
        sheet_frame.columnconfigure(0, weight=1)

        sheet_header = ttk.Frame(sheet_frame, style="Card.TFrame")
        sheet_header.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        sheet_header.columnconfigure(0, weight=1)
        ttk.Label(sheet_header, text="복사할 시트", style="SectionTitle.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        self.common_sheet_btn = ttk.Button(
            sheet_header,
            text="공통 시트 선택",
            command=self._open_common_sheet_dialog,
        )
        self.common_sheet_btn.grid(row=0, column=1, sticky="e")

        ttk.Label(
            sheet_frame,
            text="체크된 시트만 추가됩니다.",
            style="CardHint.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(0, 4))

        self.sheet_check_frame = ttk.Frame(sheet_frame, style="Inset.TFrame", padding=(8, 6, 8, 6))
        self.sheet_check_frame.grid(row=2, column=0, sticky="ew")
        self.sheet_check_frame.columnconfigure(0, weight=1)
        self.sheet_no_file_label = ttk.Label(
            self.sheet_check_frame,
            text="파일을 선택하면 시트가 표시됩니다.",
            style="CardHint.TLabel",
        )
        self.sheet_no_file_label.grid(row=0, column=0, sticky="w")
        self.sheet_check_vars: Dict[str, tk.BooleanVar] = {}

        opt_frame = ttk.LabelFrame(left, text="3. 읽기 옵션", style="Card.TLabelframe", padding=(12, 10))
        opt_frame.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        opt_frame.columnconfigure(1, weight=1)
        ttk.Label(opt_frame, text="헤더 행", style="SectionTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(opt_frame, from_=1, to=9999, textvariable=self.header_row_var, width=8).grid(
            row=0, column=1, sticky="w", padx=(10, 0)
        )
        ttk.Checkbutton(opt_frame, text="빈 행 제외", variable=self.skip_empty_rows_var).grid(
            row=1, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )

        s3 = ttk.LabelFrame(right, text="미리보기 및 규칙", style="Card.TLabelframe", padding=(16, 14))
        s3.grid(row=0, column=0, sticky="nsew", pady=(0, 12))
        s3.columnconfigure(0, weight=1)
        s3.rowconfigure(2, weight=1)
        ttk.Label(
            s3,
            text="파일을 선택하면 상단 행을 미리 볼 수 있습니다. 열 헤더를 우클릭해서 검증 규칙을 추가하세요.",
            style="CardHint.TLabel",
            wraplength=760,
            justify="left",
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(s3, textvariable=self.preview_status_var, style="CardHint.TLabel").grid(
            row=1, column=0, sticky="w", pady=(6, 10)
        )

        preview_inner = ttk.Frame(s3, style="Inset.TFrame", padding=(0, 0, 0, 0))
        preview_inner.grid(row=2, column=0, sticky="nsew")
        preview_inner.columnconfigure(0, weight=1)
        preview_inner.rowconfigure(0, weight=1)
        self.preview_tree = ttk.Treeview(preview_inner, style="Data.Treeview", show="headings", height=12)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        pv_y = ttk.Scrollbar(preview_inner, orient="vertical", command=self.preview_tree.yview)
        pv_y.grid(row=0, column=1, sticky="ns")
        pv_x = ttk.Scrollbar(preview_inner, orient="horizontal", command=self.preview_tree.xview)
        pv_x.grid(row=1, column=0, sticky="ew")
        self.preview_tree.configure(yscrollcommand=pv_y.set, xscrollcommand=pv_x.set)
        self.preview_tree.bind("<Button-3>", self._on_right_click)

        rule_bar = ttk.Frame(s3, style="Card.TFrame")
        rule_bar.grid(row=3, column=0, sticky="ew", pady=(12, 8))
        rule_bar.columnconfigure(1, weight=1)
        ttk.Label(rule_bar, text="적용된 규칙", style="SectionTitle.TLabel").grid(row=0, column=0, sticky="nw")
        self.rule_label_var = tk.StringVar(value="(없음)")
        ttk.Label(
            rule_bar,
            textvariable=self.rule_label_var,
            style="CardHint.TLabel",
            wraplength=760,
            justify="left",
        ).grid(row=0, column=1, sticky="ew", padx=(10, 0))

        rule_btn_bar = ttk.Frame(s3, style="Card.TFrame")
        rule_btn_bar.grid(row=4, column=0, sticky="ew")
        ttk.Button(rule_btn_bar, text="이 파일 규칙 삭제", command=self._clear_current_rules).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(rule_btn_bar, text="모든 파일 규칙 삭제", command=self._clear_all_rules).pack(
            side="left", padx=(0, 12)
        )
        ttk.Separator(rule_btn_bar, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Button(rule_btn_bar, text="규칙 저장", command=self._save_rules).pack(side="left", padx=(0, 6))
        ttk.Button(
            rule_btn_bar,
            text="규칙 불러오기 (이 파일)",
            command=lambda: self._load_rules(all_files=False),
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            rule_btn_bar,
            text="규칙 불러오기 (모든 파일)",
            command=lambda: self._load_rules(all_files=True),
        ).pack(side="left")

        self.ctx_menu = tk.Menu(self.root, tearoff=0)

        bottom_row = ttk.Frame(right, style="App.TFrame")
        bottom_row.grid(row=1, column=0, sticky="nsew")
        bottom_row.columnconfigure(0, weight=1)
        bottom_row.columnconfigure(1, weight=0, minsize=220)
        bottom_row.rowconfigure(0, weight=1)

        log_frame = ttk.LabelFrame(bottom_row, text="실행 로그", style="Card.TLabelframe", padding=(16, 14))
        log_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = ScrolledText(log_frame, height=8, wrap="word", padx=12, pady=12)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        self.log_text.configure(state="disabled")

        run_frame = ttk.LabelFrame(bottom_row, text="병합 실행", style="Card.TLabelframe", padding=(16, 14))
        run_frame.grid(row=0, column=1, sticky="nsew")
        run_frame.columnconfigure(0, weight=1)
        run_frame.rowconfigure(1, weight=1)
        ttk.Label(
            run_frame,
            text="준비가 끝나면 병합을 실행합니다.",
            style="CardHint.TLabel",
            wraplength=180,
            justify="left",
        ).grid(row=0, column=0, sticky="nw", pady=(0, 10))
        ttk.Frame(run_frame, style="Card.TFrame").grid(row=1, column=0, sticky="nsew")
        ttk.Button(
            run_frame,
            text="병합 실행",
            command=self.run_merge,
            style="Accent.TButton",
        ).grid(row=2, column=0, sticky="sew")

    # ── 로그 ──
    def log(self, message: str, tag: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{ts}] {message}\n", tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    # ── STEP 1 ──
    def select_target(self):
        path = filedialog.askopenfilename(title="대상 엑셀 파일 선택",
            filetypes=[("Excel File", "*.xlsx *.xlsm"), ("All Files", "*.*")])
        if path:
            self.target_path_var.set(path)
            self.log(f"대상 파일: {os.path.basename(path)}")

    def create_new_target(self):
        path = filedialog.asksaveasfilename(title="새 엑셀 파일 만들기",
            defaultextension=".xlsx", filetypes=[("Excel File", "*.xlsx"), ("All Files", "*.*")])
        if path:
            self.target_path_var.set(path)
            self.log(f"새 파일 생성 예정: {os.path.basename(path)}")

    def _file_label(self, path: str) -> str:
        rules = self.file_rules.get(path, {})
        name = os.path.basename(path)
        return f"{name}  [{len(rules)}개 규칙]" if rules else name

    def _get_common_sheet_names(self) -> List[str]:
        if not self.selected_files:
            return []
        common = None
        for path in self.selected_files:
            available = self.file_available_sheets.get(path, [])
            available_set = set(available)
            common = available_set if common is None else common & available_set
        if not common:
            return []
        return sorted(common)

    def _refresh_bulk_sheet_controls(self):
        common_sheets = self._get_common_sheet_names()
        if common_sheets:
            self.common_sheet_btn.state(["!disabled"])
        else:
            self.common_sheet_btn.state(["disabled"])

    def _open_common_sheet_dialog(self):
        common_sheets = self._get_common_sheet_names()
        if not common_sheets:
            messagebox.showwarning("확인", "공통 시트가 없습니다.")
            return

        dialog = _CommonSheetSelectionDialog(self.root, common_sheets)
        if dialog.result is None:
            return

        selected_common = dialog.result
        if not selected_common:
            messagebox.showwarning("확인", "선택된 시트가 없습니다.")
            return

        applied = 0
        for path in self.selected_files:
            available = self.file_available_sheets.get(path, [])
            matching = [s for s in available if s in selected_common]
            if matching:
                self.file_sheets[path] = matching
                applied += 1

        names = ", ".join(selected_common)
        self.log(f"모든 파일에 공통 시트 적용: [{names}] ({applied}개 파일)", "ok")
        if self.current_preview_file:
            self._update_sheet_checkboxes(self.current_preview_file)
            cur_selected = self.file_sheets.get(self.current_preview_file, [])
            if cur_selected:
                self._show_file_preview(self.current_preview_file, cur_selected[0])
            else:
                self.preview_tree.delete(*self.preview_tree.get_children())
                self.preview_tree["columns"] = []
                self.preview_columns = []
                self.preview_display_columns = []
                self.preview_status_var.set(f"{os.path.basename(self.current_preview_file)} | 선택된 시트 없음")
                self._refresh_rule_display()

    def _get_selected_file_indices(self) -> List[int]:
        indices = []
        for item_id in self.file_listbox.selection():
            try:
                idx = int(item_id)
            except (TypeError, ValueError):
                continue
            if 0 <= idx < len(self.selected_files):
                indices.append(idx)
        return sorted(indices)

    def _get_preview_column_display_label(self, column_name: str, with_position: bool = False) -> str:
        if column_name not in self.preview_columns:
            return column_name
        idx = self.preview_columns.index(column_name)
        display = self.preview_display_columns[idx] if idx < len(self.preview_display_columns) else column_name
        if with_position and self.preview_display_columns.count(display) > 1:
            return f"{display} ({idx + 1}번째 열)"
        return display

    def _refresh_file_listbox(self, selected_path: Optional[str] = None):
        if selected_path is None:
            selection = self._get_selected_file_indices()
            if selection:
                idx = selection[0]
                selected_path = self.selected_files[idx]

        children = self.file_listbox.get_children()
        if children:
            self.file_listbox.delete(*children)
        for idx, path in enumerate(self.selected_files):
            self.file_listbox.insert("", "end", iid=str(idx), values=(self._file_label(path),))

        if selected_path in self.selected_files:
            idx = self.selected_files.index(selected_path)
            item_id = str(idx)
            self.file_listbox.selection_set(item_id)
            self.file_listbox.focus(item_id)
            self.file_listbox.see(item_id)

    # ── STEP 2 ──
    def add_files(self):
        paths = filedialog.askopenfilenames(title="추가할 엑셀 파일 선택",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All Files", "*.*")])
        if not paths:
            return
        added = 0
        focus_path = None
        skipped_duplicates = []
        skipped_unsupported = []
        for path in paths:
            ext = os.path.splitext(path)[1].lower()
            if ext not in SUPPORTED_EXTENSIONS:
                skipped_unsupported.append(os.path.basename(path))
                continue
            if path not in self.selected_files:
                self.selected_files.append(path)
                # 시트 목록 로드 - 기본으로 전체 선택
                try:
                    wb = load_workbook(filename=path, read_only=True)
                    sheet_names = list(wb.sheetnames)
                    self.file_available_sheets[path] = sheet_names
                    self.file_sheets[path] = list(sheet_names)
                    wb.close()
                except Exception:
                    self.file_available_sheets[path] = []
                    self.file_sheets[path] = []
                if focus_path is None:
                    focus_path = path
                added += 1
            else:
                skipped_duplicates.append(os.path.basename(path))
        if added:
            self.log(f"파일 {added}개 추가됨")
            self._refresh_bulk_sheet_controls()
            focus_path = focus_path or self.selected_files[0]
            self._refresh_file_listbox(selected_path=focus_path)
            self._update_sheet_checkboxes(focus_path)
            self._show_file_preview(focus_path)
        else:
            reasons = []
            if skipped_duplicates:
                reasons.append(f"이미 추가된 파일: {', '.join(skipped_duplicates[:3])}")
            if skipped_unsupported:
                reasons.append(f"지원하지 않는 형식: {', '.join(skipped_unsupported[:3])}")
            detail = "\n".join(reasons) if reasons else "파일이 추가되지 않았습니다."
            messagebox.showwarning("추가 실패", detail)

        if skipped_duplicates:
            self.log(f"중복 파일 건너뜀: {len(skipped_duplicates)}개", "warn")
        if skipped_unsupported:
            self.log(f"지원하지 않는 형식 건너뜀: {len(skipped_unsupported)}개", "warn")

    def remove_selected_files(self):
        selection = self._get_selected_file_indices()
        if not selection:
            return
        next_index = min(selection[0], len(self.selected_files) - len(selection) - 1)
        for index in reversed(selection):
            path = self.selected_files[index]
            self.file_rules.pop(path, None)
            self.file_sheets.pop(path, None)
            self.file_available_sheets.pop(path, None)
            del self.selected_files[index]
        self._refresh_bulk_sheet_controls()
        self._refresh_file_listbox()
        if self.selected_files:
            next_index = max(0, min(next_index, len(self.selected_files) - 1))
            next_path = self.selected_files[next_index]
            self._refresh_file_listbox(selected_path=next_path)
            self._update_sheet_checkboxes(next_path)
            self._show_file_preview(next_path)
        else:
            self._clear_preview()

    def clear_files(self):
        self.selected_files.clear()
        self.file_rules.clear()
        self.file_sheets.clear()
        self.file_available_sheets.clear()
        self._refresh_bulk_sheet_controls()
        self._refresh_file_listbox()
        self._clear_preview()

    def _on_file_select(self, _event=None):
        selection = self._get_selected_file_indices()
        if not selection:
            return
        idx = selection[0]
        if 0 <= idx < len(self.selected_files):
            path = self.selected_files[idx]
            self._update_sheet_checkboxes(path)
            self._show_file_preview(path)

    def _update_sheet_checkboxes(self, file_path: str):
        """파일의 시트 목록을 체크박스로 표시"""
        self.sheet_check_vars.clear()
        for w in self._sheet_check_widgets:
            w.destroy()
        self._sheet_check_widgets.clear()
        self.sheet_no_file_label.grid_remove()

        all_sheets = self.file_available_sheets.get(file_path, [])
        selected = self.file_sheets.get(file_path, all_sheets)

        if not all_sheets:
            self.sheet_no_file_label.configure(text="시트 없음")
            self.sheet_no_file_label.grid()
            return

        for idx, sn in enumerate(all_sheets):
            var = tk.BooleanVar(value=(sn in selected))
            self.sheet_check_vars[sn] = var
            cb = ttk.Checkbutton(
                self.sheet_check_frame,
                text=sn,
                variable=var,
                style="Panel.TCheckbutton",
                command=lambda fp=file_path: self._on_sheet_toggle(fp),
            )
            cb.grid(row=1 + idx // 2, column=idx % 2, sticky="w", padx=(0, 16), pady=(0, 4))
            self._sheet_check_widgets.append(cb)

    def _on_sheet_toggle(self, file_path: str):
        """체크박스 변경 시 선택된 시트 목록 업데이트"""
        selected = [sn for sn, var in self.sheet_check_vars.items() if var.get()]
        self.file_sheets[file_path] = selected
        self._refresh_bulk_sheet_controls()
        if file_path == self.current_preview_file:
            if selected:
                self._show_file_preview(file_path, selected[0])
            else:
                self.preview_tree.delete(*self.preview_tree.get_children())
                self.preview_tree["columns"] = []
                self.preview_columns = []
                self.preview_display_columns = []
                self.preview_status_var.set(f"{os.path.basename(file_path)} | 선택된 시트 없음")
                self._refresh_rule_display()

    # ── 미리보기 ──
    def _show_file_preview(self, file_path: str, sheet_name: str = None):
        self.current_preview_file = file_path
        rules = self.file_rules.get(file_path, {})
        try:
            wb = load_workbook(filename=file_path, data_only=False, read_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # 선택된 시트 중 첫 번째로 미리보기
                selected = self.file_sheets.get(file_path)
                if selected is None:
                    selected = list(wb.sheetnames)
                if not selected:
                    wb.close()
                    self.preview_tree.delete(*self.preview_tree.get_children())
                    self.preview_tree["columns"] = []
                    self.preview_columns = []
                    self.preview_display_columns = []
                    self.preview_status_var.set(f"{os.path.basename(file_path)} | 선택된 시트 없음")
                    self._refresh_rule_display()
                    return
                ws = wb[selected[0]]
            hr = self.header_row_var.get()
            mc = min(ws.max_column or 1, 30)
            mr = min(ws.max_row or 1, 15)
            rows_data = list(ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc, values_only=True))
            wb.close()

            if not rows_data or hr < 1 or hr > len(rows_data):
                self.preview_tree.delete(*self.preview_tree.get_children())
                self.preview_tree["columns"] = []
                self.preview_columns = []
                self.preview_display_columns = []
                self.preview_status_var.set(
                    f"{os.path.basename(file_path)} | 시트: {ws.title} | 헤더 행 {hr} 미리보기 불가"
                )
                self._refresh_rule_display()
                return

            header_values = list(rows_data[hr - 1])
            columns = normalize_headers(header_values)
            display_columns = []
            for i, value in enumerate(header_values, start=1):
                name = str(value).strip() if value is not None else ""
                display_columns.append(name if name else f"Column_{i}")

            self.preview_columns = columns
            self.preview_display_columns = display_columns
            self.preview_tree.delete(*self.preview_tree.get_children())
            self.preview_tree["columns"] = columns
            for idx, col in enumerate(columns):
                display = display_columns[idx]
                heading_text = f"{display}  [{rules[col].rule_type}]" if col in rules else display
                self.preview_tree.heading(col, text=heading_text)
                self.preview_tree.column(col, width=130, anchor="w", stretch=False)

            for row_data in rows_data[hr:]:
                vals = []
                for v in row_data:
                    if v is None:
                        vals.append("")
                    else:
                        t = str(v)
                        vals.append(t[:50] + "..." if len(t) > 50 else t)
                self.preview_tree.insert("", "end", values=vals)

            self.preview_status_var.set(
                f"{os.path.basename(file_path)} | 시트: {ws.title} | 헤더 행 {hr}"
            )
            self._refresh_rule_display()
        except Exception as exc:
            self.preview_tree.delete(*self.preview_tree.get_children())
            self.preview_tree["columns"] = []
            self.preview_columns = []
            self.preview_display_columns = []
            self.preview_status_var.set("미리보기 로드 실패")
            self._refresh_rule_display()
            self.log(f"미리보기 실패: {exc}", "error")

    def _clear_preview(self):
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = []
        self.preview_columns = []
        self.preview_display_columns = []
        self.current_preview_file = None
        self.preview_status_var.set("선택된 파일 없음")
        self.rule_label_var.set("(없음)")
        self.sheet_check_vars.clear()
        for w in self._sheet_check_widgets:
            w.destroy()
        self._sheet_check_widgets.clear()
        self.sheet_no_file_label.configure(text="파일을 선택하면 시트가 표시됩니다.")
        self.sheet_no_file_label.grid()

    # ── 우클릭 메뉴 ──
    def _on_right_click(self, event):
        if not self.current_preview_file:
            return
        region = self.preview_tree.identify_region(event.x, event.y)
        if region != "heading":
            return
        col_id = self.preview_tree.identify_column(event.x)
        try:
            ci = int(col_id.replace("#", "")) - 1
        except (ValueError, TypeError):
            return
        if ci < 0 or ci >= len(self.preview_columns):
            return
        col_name = self.preview_columns[ci]
        display_col_name = self._get_preview_column_display_label(col_name, with_position=True)
        rules = self.file_rules.get(self.current_preview_file, {})

        self.ctx_menu.delete(0, "end")
        self.ctx_menu.add_command(
            label=f"열 검증 규칙: {display_col_name}",
            command=lambda: None,
            foreground=self.palette["muted"],
            activebackground=self.palette["surface"],
            activeforeground=self.palette["text"],
        )
        self.ctx_menu.add_separator()

        # ── 이 파일에만 ──
        this_menu = tk.Menu(self.ctx_menu, tearoff=0)
        this_menu.configure(
            bg=self.palette["surface"],
            fg=self.palette["text"],
            activebackground=self.palette["selection"],
            activeforeground=self.palette["selection_text"],
            relief="flat",
            borderwidth=1,
        )
        for rt in SIMPLE_RULE_TYPES:
            this_menu.add_command(label=rt,
                command=lambda r=rt, c=col_name: self._add_rule_this(c, r))
        this_menu.add_separator()
        this_menu.add_command(label="숫자 범위...",
            command=lambda c=col_name: self._add_range_this(c))
        this_menu.add_command(label="텍스트 포함...",
            command=lambda c=col_name: self._add_text_this(c))
        this_menu.add_command(label="허용값 목록...",
            command=lambda c=col_name: self._add_allowed_this(c))
        this_menu.add_separator()
        this_menu.add_command(label="조건부 검증...",
            command=lambda c=col_name: self._add_conditional_this(c))
        self.ctx_menu.add_cascade(label="이 파일에만 적용", menu=this_menu)

        # ── 모든 파일에 ──
        all_menu = tk.Menu(self.ctx_menu, tearoff=0)
        all_menu.configure(
            bg=self.palette["surface"],
            fg=self.palette["text"],
            activebackground=self.palette["selection"],
            activeforeground=self.palette["selection_text"],
            relief="flat",
            borderwidth=1,
        )
        for rt in SIMPLE_RULE_TYPES:
            all_menu.add_command(label=rt,
                command=lambda r=rt, c=col_name: self._add_rule_all(c, r))
        all_menu.add_separator()
        all_menu.add_command(label="숫자 범위...",
            command=lambda c=col_name: self._add_range_all(c))
        all_menu.add_command(label="텍스트 포함...",
            command=lambda c=col_name: self._add_text_all(c))
        all_menu.add_command(label="허용값 목록...",
            command=lambda c=col_name: self._add_allowed_all(c))
        all_menu.add_separator()
        all_menu.add_command(label="조건부 검증...",
            command=lambda c=col_name: self._add_conditional_all(c))
        self.ctx_menu.add_cascade(label="모든 파일에 적용", menu=all_menu)

        if col_name in rules:
            self.ctx_menu.add_separator()
            self.ctx_menu.add_command(label=f"{display_col_name} 규칙 삭제 (이 파일)",
                command=lambda c=col_name: self._remove_rule_this(c))
            self.ctx_menu.add_command(label=f"{display_col_name} 규칙 삭제 (모든 파일)",
                command=lambda c=col_name: self._remove_rule_all(c))

        self.ctx_menu.post(event.x_root, event.y_root)

    # ── 규칙 추가 (이 파일) ──
    def _set_rule(self, file_path: str, rule: ValidationRule):
        if file_path not in self.file_rules:
            self.file_rules[file_path] = {}
        self.file_rules[file_path][rule.column] = rule

    def _add_rule_this(self, col: str, rule_type: str):
        d = _SimpleMsgDialog(self.root, col, rule_type)
        if d.result is not None:
            self._set_rule(self.current_preview_file,
                ValidationRule(column=col, rule_type=rule_type, error_msg=d.result))
            self.log(f"[{os.path.basename(self.current_preview_file)}] [{col}] {rule_type}", "ok")
            self._refresh_after_rule_change()

    def _add_range_this(self, col: str):
        d = _RangeDialog(self.root, col)
        if d.result:
            self._set_rule(self.current_preview_file,
                ValidationRule(column=col, rule_type="숫자 범위",
                               value1=d.result[0], value2=d.result[1], error_msg=d.result[2]))
            self.log(f"[{os.path.basename(self.current_preview_file)}] [{col}] 숫자 범위", "ok")
            self._refresh_after_rule_change()

    def _add_text_this(self, col: str):
        d = _ValueMsgDialog(self.root, "텍스트 포함",
                            f"'{col}' 열에 포함되어야 할 텍스트:")
        if d.result:
            self._set_rule(self.current_preview_file,
                ValidationRule(column=col, rule_type="텍스트 포함",
                               value1=d.result[0], error_msg=d.result[1]))
            self.log(f"[{os.path.basename(self.current_preview_file)}] [{col}] '{d.result[0]}' 포함", "ok")
            self._refresh_after_rule_change()

    def _add_allowed_this(self, col: str):
        d = _ValueMsgDialog(self.root, "허용값 목록",
                            f"'{col}' 열 허용값 (쉼표 구분):",
                            hint="예: 합격, 불합격, 보류")
        if d.result:
            self._set_rule(self.current_preview_file,
                ValidationRule(column=col, rule_type="허용값 목록",
                               value1=d.result[0], error_msg=d.result[1]))
            self.log(f"[{os.path.basename(self.current_preview_file)}] [{col}] 허용값({d.result[0]})", "ok")
            self._refresh_after_rule_change()

    # ── 규칙 추가 (모든 파일) ──
    def _add_rule_all(self, col: str, rule_type: str):
        d = _SimpleMsgDialog(self.root, col, rule_type)
        if d.result is not None:
            for f in self.selected_files:
                self._set_rule(f, ValidationRule(column=col, rule_type=rule_type, error_msg=d.result))
            self.log(f"[모든 파일] [{col}] {rule_type}", "ok")
            self._refresh_after_rule_change()

    def _add_range_all(self, col: str):
        d = _RangeDialog(self.root, col)
        if d.result:
            for f in self.selected_files:
                self._set_rule(f, ValidationRule(column=col, rule_type="숫자 범위",
                               value1=d.result[0], value2=d.result[1], error_msg=d.result[2]))
            self.log(f"[모든 파일] [{col}] 숫자 범위", "ok")
            self._refresh_after_rule_change()

    def _add_text_all(self, col: str):
        d = _ValueMsgDialog(self.root, "텍스트 포함",
                            f"'{col}' 열에 포함되어야 할 텍스트:")
        if d.result:
            for f in self.selected_files:
                self._set_rule(f, ValidationRule(column=col, rule_type="텍스트 포함",
                               value1=d.result[0], error_msg=d.result[1]))
            self.log(f"[모든 파일] [{col}] '{d.result[0]}' 포함", "ok")
            self._refresh_after_rule_change()

    def _add_allowed_all(self, col: str):
        d = _ValueMsgDialog(self.root, "허용값 목록",
                            f"'{col}' 열 허용값 (쉼표 구분):",
                            hint="예: 합격, 불합격, 보류")
        if d.result:
            for f in self.selected_files:
                self._set_rule(f, ValidationRule(column=col, rule_type="허용값 목록",
                               value1=d.result[0], error_msg=d.result[1]))
            self.log(f"[모든 파일] [{col}] 허용값({d.result[0]})", "ok")
            self._refresh_after_rule_change()

    # ── 조건부 검증 ──
    def _add_conditional_this(self, col: str):
        d = _ConditionalDialog(self.root, col, self.preview_columns)
        if d.result:
            self._set_rule(self.current_preview_file,
                ValidationRule(column=col, rule_type="조건부 검증",
                               value1=d.result[0], value2=d.result[1],
                               value3=d.result[2], error_msg=d.result[3]))
            self.log(f"[{os.path.basename(self.current_preview_file)}] [{col}] 조건부 검증", "ok")
            self._refresh_after_rule_change()

    def _add_conditional_all(self, col: str):
        d = _ConditionalDialog(self.root, col, self.preview_columns)
        if d.result:
            for f in self.selected_files:
                self._set_rule(f, ValidationRule(column=col, rule_type="조건부 검증",
                               value1=d.result[0], value2=d.result[1],
                               value3=d.result[2], error_msg=d.result[3]))
            self.log(f"[모든 파일] [{col}] 조건부 검증", "ok")
            self._refresh_after_rule_change()

    # ── 규칙 삭제 ──
    def _remove_rule_this(self, col: str):
        if self.current_preview_file and self.current_preview_file in self.file_rules:
            self.file_rules[self.current_preview_file].pop(col, None)
        self.log(f"규칙 삭제: [{col}] (이 파일)")
        self._refresh_after_rule_change()

    def _remove_rule_all(self, col: str):
        for rules in self.file_rules.values():
            rules.pop(col, None)
        self.log(f"규칙 삭제: [{col}] (모든 파일)")
        self._refresh_after_rule_change()

    def _clear_current_rules(self):
        if self.current_preview_file:
            self.file_rules.pop(self.current_preview_file, None)
            self.log(f"규칙 삭제: {os.path.basename(self.current_preview_file)}")
            self._refresh_after_rule_change()

    def _clear_all_rules(self):
        self.file_rules.clear()
        self.log("모든 파일의 규칙 삭제됨")
        self._refresh_after_rule_change()

    def _refresh_after_rule_change(self):
        self._refresh_rule_display()
        self._refresh_preview_headers()
        if not self.current_preview_file:
            self._refresh_file_listbox()

    def _refresh_rule_display(self):
        if not self.current_preview_file:
            self.rule_label_var.set("(없음)")
            return
        rules = self.file_rules.get(self.current_preview_file, {})
        if not rules:
            self.rule_label_var.set("없음. 열 헤더를 우클릭해 규칙을 추가하세요.")
            return
        parts = [f"[{self._get_preview_column_display_label(col, with_position=True)}] {r.display()}" for col, r in rules.items()]
        self.rule_label_var.set("  |  ".join(parts))

    def _refresh_preview_headers(self):
        if not self.current_preview_file:
            return
        rules = self.file_rules.get(self.current_preview_file, {})
        for idx, col in enumerate(self.preview_columns):
            display = self.preview_display_columns[idx] if idx < len(self.preview_display_columns) else col
            heading_text = f"{display}  [{rules[col].rule_type}]" if col in rules else display
            self.preview_tree.heading(col, text=heading_text)

        self._refresh_file_listbox(selected_path=self.current_preview_file)

    # ── 규칙 저장/불러오기 ──
    def _save_rules(self):
        if not self.current_preview_file:
            messagebox.showwarning("확인", "파일을 먼저 선택하세요.")
            return
        rules = self.file_rules.get(self.current_preview_file, {})
        if not rules:
            messagebox.showwarning("확인", "저장할 규칙이 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            title="검증 규칙 저장",
            defaultextension=".json",
            filetypes=[("JSON 파일", "*.json"), ("All Files", "*.*")],
        )
        if not path:
            return
        data = [r.to_dict() for r in rules.values()]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.log(f"규칙 저장: {os.path.basename(path)} ({len(data)}개)", "ok")

    def _load_rules(self, all_files: bool):
        path = filedialog.askopenfilename(
            title="검증 규칙 불러오기",
            filetypes=[("JSON 파일", "*.json"), ("All Files", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            loaded = {}
            for d in data:
                rule = ValidationRule.from_dict(d)
                loaded[rule.column] = rule
        except Exception as exc:
            messagebox.showerror("오류", f"규칙 불러오기 실패:\n{exc}")
            return

        if all_files:
            for fp in self.selected_files:
                if fp not in self.file_rules:
                    self.file_rules[fp] = {}
                self.file_rules[fp].update({col: ValidationRule.from_dict(r.to_dict()) for col, r in loaded.items()})
            self.log(f"규칙 불러오기: {os.path.basename(path)} -> 모든 파일 ({len(loaded)}개)", "ok")
        else:
            if self.current_preview_file:
                if self.current_preview_file not in self.file_rules:
                    self.file_rules[self.current_preview_file] = {}
                self.file_rules[self.current_preview_file].update(loaded)
                self.log(f"규칙 불러오기: {os.path.basename(path)} -> {os.path.basename(self.current_preview_file)} ({len(loaded)}개)", "ok")
            else:
                messagebox.showwarning("확인", "파일을 먼저 선택하세요.")
                return

        self._refresh_after_rule_change()

    # ── 시트 복사 ──
    def _copy_sheet(self, source_ws, target_wb, new_name: str):
        target_ws = target_wb.create_sheet(title=new_name)
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))
        for cl, dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[cl].width = dim.width
        for rn, dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[rn].height = dim.height

    # ── 병합 실행 ──
    def run_merge(self):
        if not self.selected_files:
            messagebox.showwarning("확인", "추가할 파일을 먼저 선택하세요.")
            return
        target_path = self.target_path_var.get().strip()
        if not target_path:
            messagebox.showwarning("확인", "대상 파일을 지정하세요.")
            return
        if os.path.splitext(target_path)[1].lower() not in {".xlsx", ".xlsm"}:
            messagebox.showerror("오류", ".xlsx 또는 .xlsm 파일만 지원합니다.")
            return

        append_mode = os.path.exists(target_path)
        if not append_mode:
            d = os.path.dirname(os.path.abspath(target_path))
            if d:
                os.makedirs(d, exist_ok=True)

        hr = self.header_row_var.get()
        se = bool(self.skip_empty_rows_var.get())

        all_empty, all_error, all_rule = [], [], []
        failed, added_sheets = [], []

        self.log(f"병합 시작: {len(self.selected_files)}개 파일 -> {os.path.basename(target_path)}")

        try:
            if append_mode:
                twb = load_workbook(filename=target_path)
            else:
                twb = Workbook()
                if twb.sheetnames:
                    twb.remove(twb.active)
            existing = set(twb.sheetnames)

            total_issues = 0
            from openpyxl.styles import Font, PatternFill, Alignment
            for path in self.selected_files:
                fname = os.path.basename(path)
                base_name = os.path.splitext(fname)[0]
                try:
                    swb = load_workbook(filename=path)
                    # 선택된 시트만 복사 (선택 정보 없으면 전체)
                    selected_sheets = self.file_sheets.get(path, list(swb.sheetnames))
                    sheets_to_copy = [s for s in swb.sheetnames if s in selected_sheets]
                    if not sheets_to_copy:
                        swb.close()
                        self.log(f"  {fname}: 선택된 시트 없음 - 건너뜀", "error")
                        continue

                    for src_sheet_name in sheets_to_copy:
                        sws = swb[src_sheet_name]
                        name_candidate = src_sheet_name
                        sheet_name = unique_sheet_name(name_candidate, existing)
                        existing.add(sheet_name)
                        self._copy_sheet(sws, twb, sheet_name)
                        added_sheets.append(sheet_name)

                        # 검증: 행별 오류 메시지 수집
                        rules = self.file_rules.get(path, {})
                        df, actual = read_excel_data(path, src_sheet_name, hr, se)
                        row_msgs = validate_rules_by_row(df, rules)

                        # 복사된 시트에 비고 열 추가
                        target_ws = twb[sheet_name]
                        bigo_col = (target_ws.max_column or 1) + 1

                        # 헤더 행에 "비고" 제목
                        header_cell = target_ws.cell(row=hr, column=bigo_col, value="비고")
                        header_cell.font = Font(bold=True, color="FFFFFF")
                        header_cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                        header_cell.alignment = Alignment(horizontal="center")
                        target_ws.column_dimensions[header_cell.column_letter].width = 40

                        issue_count = 0
                        for src_row, messages in row_msgs.items():
                            cell = target_ws.cell(row=src_row, column=bigo_col,
                                                  value=" / ".join(messages))
                            cell.font = Font(color="CC0000")
                            issue_count += len(messages)
                        total_issues += issue_count

                        rc = sws.max_row or 0
                        self.log(f"  {fname} -> [{sheet_name}] {rc}행 | 문제 {issue_count}건")

                    swb.close()
                except Exception as exc:
                    failed.append((fname, str(exc)))
                    self.log(f"  {fname}: 실패 - {exc}", "error")

            if not added_sheets:
                twb.close()
                messagebox.showerror("실패", "처리된 파일이 없습니다.")
                return
            twb.save(target_path)
            twb.close()

        except PermissionError:
            self.log("저장 실패: 파일이 열려 있습니다.", "error")
            messagebox.showerror("저장 실패",
                f"'{os.path.basename(target_path)}' 파일이 열려 있습니다.\n파일을 닫고 다시 시도하세요.")
            return
        except Exception as exc:
            self.log(f"저장 실패: {exc}", "error")
            messagebox.showerror("저장 실패", str(exc))
            return

        self.log(f"병합 완료! 시트 {len(added_sheets)}개 추가: {', '.join(added_sheets)}", "ok")
        if total_issues:
            self.log(f"검증 문제 총 {total_issues}건 → 각 시트 '비고' 열에 기록됨", "warn")
        else:
            self.log("검증 문제 없음", "ok")

        msg = f"병합 완료!\n\n대상: {os.path.basename(target_path)}\n추가된 시트: {', '.join(added_sheets)}\n"
        if total_issues:
            msg += f"\n검증 문제 {total_issues}건이 각 시트의\n'비고' 열에 기록되었습니다."
        else:
            msg += "\n검증 문제 없음"
        if failed:
            msg += f"\n\n실패: {len(failed)}개 파일"
        messagebox.showinfo("완료", msg)


class _CommonSheetSelectionDialog(tk.Toplevel):
    """모든 파일에 공통으로 존재하는 시트를 선택하여 일괄 적용하는 다이얼로그"""

    def __init__(self, parent, common_sheets: List[str]):
        super().__init__(parent)
        apply_window_palette(self, getattr(parent, "_excel_palette", get_palette("light")))
        self.title("공통 시트 선택")
        self.resizable(False, False)
        self.result = None
        self.transient(parent)
        self.grab_set()

        f = ttk.Frame(self, padding=20)
        f.pack(fill="both", expand=True)
        ttk.Label(f, text="공통 시트 선택", font=("", 10, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Label(f, text="선택한 시트가 모든 파일에 일괄 적용됩니다.", style="Hint.TLabel").pack(anchor="w", pady=(0, 10))

        list_frame = ttk.Frame(f, style="Inset.TFrame", padding=(10, 10, 10, 10))
        list_frame.pack(fill="both", expand=True)

        self.sheet_vars: Dict[str, tk.BooleanVar] = {}
        for idx, sheet_name in enumerate(common_sheets):
            var = tk.BooleanVar(value=True)
            self.sheet_vars[sheet_name] = var
            ttk.Checkbutton(
                list_frame,
                text=sheet_name,
                variable=var,
                style="Panel.TCheckbutton",
            ).grid(row=idx // 2, column=idx % 2, sticky="w", padx=(0, 16), pady=(0, 8))

        action_row = ttk.Frame(f)
        action_row.pack(fill="x", pady=(12, 0))
        ttk.Button(action_row, text="전체 선택", command=self._select_all, width=10).pack(side="left", padx=(0, 6))
        ttk.Button(action_row, text="전체 해제", command=self._clear_all, width=10).pack(side="left")

        bf = ttk.Frame(f)
        bf.pack(pady=(12, 0))
        ttk.Button(bf, text="적용", command=self._ok, width=10).pack(side="left", padx=4)
        ttk.Button(bf, text="취소", command=self.destroy, width=10).pack(side="left", padx=4)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _select_all(self):
        for var in self.sheet_vars.values():
            var.set(True)

    def _clear_all(self):
        for var in self.sheet_vars.values():
            var.set(False)

    def _ok(self):
        self.result = [name for name, var in self.sheet_vars.items() if var.get()]
        self.destroy()


class _RangeDialog(tk.Toplevel):
    def __init__(self, parent, col_name: str):
        super().__init__(parent)
        apply_window_palette(self, getattr(parent, "_excel_palette", get_palette("light")))
        self.title("숫자 범위 규칙")
        self.resizable(False, False)
        self.result = None
        self.transient(parent)
        self.grab_set()

        f = ttk.Frame(self, padding=20)
        f.pack()
        ttk.Label(f, text=f"{col_name} 열의 숫자 범위", font=("", 10)).pack(pady=(0, 12))

        r1 = ttk.Frame(f)
        r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="최소값:", width=8).pack(side="left")
        self.min_e = ttk.Entry(r1, width=15)
        self.min_e.pack(side="left", padx=4)
        ttk.Label(r1, text="(비우면 제한 없음)", style="Hint.TLabel").pack(side="left")

        r2 = ttk.Frame(f)
        r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="최대값:", width=8).pack(side="left")
        self.max_e = ttk.Entry(r2, width=15)
        self.max_e.pack(side="left", padx=4)
        ttk.Label(r2, text="(비우면 제한 없음)", style="Hint.TLabel").pack(side="left")

        r3 = ttk.Frame(f)
        r3.pack(fill="x", pady=4)
        ttk.Label(r3, text="오류 메시지:").pack(anchor="w")
        self.msg_e = ttk.Entry(r3, width=42)
        self.msg_e.pack(fill="x", pady=(4, 2))
        ttk.Label(r3, text="비워두면 자동 생성됩니다.", style="Hint.TLabel").pack(anchor="w")

        bf = ttk.Frame(f)
        bf.pack(pady=(12, 0))
        ttk.Button(bf, text="확인", command=self._ok, width=10).pack(side="left", padx=4)
        ttk.Button(bf, text="취소", command=self.destroy, width=10).pack(side="left", padx=4)

        self.min_e.focus_set()
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _ok(self):
        mn, mx = self.min_e.get().strip(), self.max_e.get().strip()
        for v, n in [(mn, "최소값"), (mx, "최대값")]:
            if v:
                try:
                    float(v)
                except ValueError:
                    messagebox.showerror("오류", f"{n}은 숫자여야 합니다.", parent=self)
                    return
        if not mn and not mx:
            messagebox.showwarning("확인", "최소값 또는 최대값을 입력하세요.", parent=self)
            return
        self.result = (mn, mx, self.msg_e.get().strip())
        self.destroy()


class _ValueMsgDialog(tk.Toplevel):
    """값 + 오류 메시지를 함께 입력받는 다이얼로그"""
    def __init__(self, parent, title: str, prompt: str, hint: str = ""):
        super().__init__(parent)
        apply_window_palette(self, getattr(parent, "_excel_palette", get_palette("light")))
        self.title(title)
        self.resizable(False, False)
        self.result = None
        self.transient(parent)
        self.grab_set()

        f = ttk.Frame(self, padding=20)
        f.pack()

        ttk.Label(f, text=prompt, font=("", 10)).pack(pady=(0, 8))
        if hint:
            ttk.Label(f, text=hint, style="Hint.TLabel").pack(pady=(0, 8))

        r1 = ttk.Frame(f)
        r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="값:", width=8).pack(side="left")
        self.val_e = ttk.Entry(r1, width=30)
        self.val_e.pack(side="left", padx=4)

        r2 = ttk.Frame(f)
        r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="오류 메시지:").pack(anchor="w")
        self.msg_e = ttk.Entry(r2, width=42)
        self.msg_e.pack(fill="x", pady=(4, 2))
        ttk.Label(r2, text="비워두면 자동 생성됩니다.", style="Hint.TLabel").pack(anchor="w")

        bf = ttk.Frame(f)
        bf.pack(pady=(12, 0))
        ttk.Button(bf, text="확인", command=self._ok, width=10).pack(side="left", padx=4)
        ttk.Button(bf, text="취소", command=self.destroy, width=10).pack(side="left", padx=4)

        self.val_e.focus_set()
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _ok(self):
        v = self.val_e.get().strip()
        if not v:
            messagebox.showwarning("확인", "값을 입력하세요.", parent=self)
            return
        self.result = (v, self.msg_e.get().strip())
        self.destroy()


class _SimpleMsgDialog(tk.Toplevel):
    """간단한 규칙의 오류 메시지만 입력받는 다이얼로그"""
    def __init__(self, parent, col_name: str, rule_type: str):
        super().__init__(parent)
        apply_window_palette(self, getattr(parent, "_excel_palette", get_palette("light")))
        self.title("오류 메시지 설정")
        self.resizable(False, False)
        self.result = None
        self.transient(parent)
        self.grab_set()

        f = ttk.Frame(self, padding=20)
        f.pack()
        ttk.Label(f, text=f"{col_name} - {rule_type}", font=("", 10)).pack(pady=(0, 8))

        r1 = ttk.Frame(f)
        r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="오류 메시지:").pack(anchor="w")
        self.msg_e = ttk.Entry(r1, width=42)
        self.msg_e.pack(fill="x", pady=(4, 2))
        ttk.Label(r1, text="비워두면 자동 생성됩니다.", style="Hint.TLabel").pack(anchor="w")

        bf = ttk.Frame(f)
        bf.pack(pady=(12, 0))
        ttk.Button(bf, text="확인", command=self._ok, width=10).pack(side="left", padx=4)
        ttk.Button(bf, text="취소", command=self.destroy, width=10).pack(side="left", padx=4)

        self.msg_e.focus_set()
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _ok(self):
        self.result = self.msg_e.get().strip()
        self.destroy()


class _ConditionalDialog(tk.Toplevel):
    """조건부 검증 다이얼로그: A열이 X이면 B열은 Y여야 한다"""
    def __init__(self, parent, target_col: str, columns: List[str]):
        super().__init__(parent)
        apply_window_palette(self, getattr(parent, "_excel_palette", get_palette("light")))
        self.title("조건부 검증 규칙")
        self.resizable(False, False)
        self.result = None
        self.transient(parent)
        self.grab_set()

        f = ttk.Frame(self, padding=20)
        f.pack()
        ttk.Label(f, text=f"{target_col} 열의 조건부 검증", font=("", 10, "bold")).pack(pady=(0, 12))
        ttk.Label(f, text="다른 열의 값이 특정 값일 때, 이 열의 값을 검증합니다.",
                  style="Hint.TLabel").pack(pady=(0, 12))

        # 조건 열 선택
        r1 = ttk.Frame(f)
        r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="조건 열:", width=10).pack(side="left")
        other_cols = [c for c in columns if c != target_col]
        self.cond_col_var = tk.StringVar(value=other_cols[0] if other_cols else "")
        cond_cb = ttk.Combobox(r1, textvariable=self.cond_col_var, values=other_cols,
                                state="readonly", width=18)
        cond_cb.pack(side="left", padx=4)

        # 조건 값
        r2 = ttk.Frame(f)
        r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="조건 값:", width=10).pack(side="left")
        self.cond_val_e = ttk.Entry(r2, width=20)
        self.cond_val_e.pack(side="left", padx=4)
        ttk.Label(r2, text="(이 값일 때 검증)", style="Hint.TLabel").pack(side="left")

        # 허용값
        r3 = ttk.Frame(f)
        r3.pack(fill="x", pady=4)
        ttk.Label(r3, text=f"'{target_col}' 허용값:", width=10).pack(side="left")
        self.allowed_e = ttk.Entry(r3, width=20)
        self.allowed_e.pack(side="left", padx=4)
        ttk.Label(r3, text="(쉼표 구분)", style="Hint.TLabel").pack(side="left")

        # 오류 메시지
        r4 = ttk.Frame(f)
        r4.pack(fill="x", pady=4)
        ttk.Label(r4, text="오류 메시지:").pack(anchor="w")
        self.msg_e = ttk.Entry(r4, width=42)
        self.msg_e.pack(fill="x", pady=(4, 2))
        ttk.Label(r4, text="비워두면 자동 생성됩니다.", style="Hint.TLabel").pack(anchor="w")

        # 미리보기 라벨
        self.preview_var = tk.StringVar()
        ttk.Label(f, textvariable=self.preview_var, style="AccentText.TLabel",
                  wraplength=350).pack(pady=(8, 0))
        self._update_preview()
        self.cond_col_var.trace_add("write", lambda *_: self._update_preview())
        self.cond_val_e.bind("<KeyRelease>", lambda _: self._update_preview())
        self.allowed_e.bind("<KeyRelease>", lambda _: self._update_preview())

        bf = ttk.Frame(f)
        bf.pack(pady=(12, 0))
        ttk.Button(bf, text="확인", command=self._ok, width=10).pack(side="left", padx=4)
        ttk.Button(bf, text="취소", command=self.destroy, width=10).pack(side="left", padx=4)

        self.cond_val_e.focus_set()
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def _update_preview(self):
        cc = self.cond_col_var.get()
        cv = self.cond_val_e.get().strip()
        av = self.allowed_e.get().strip()
        if cc and cv and av:
            self.preview_var.set(f"→ [{cc}] 값이 '{cv}'이면 허용값: {av}")
        else:
            self.preview_var.set("")

    def _ok(self):
        cc = self.cond_col_var.get().strip()
        cv = self.cond_val_e.get().strip()
        av = self.allowed_e.get().strip()
        if not cc:
            messagebox.showwarning("확인", "조건 열을 선택하세요.", parent=self)
            return
        if not cv:
            messagebox.showwarning("확인", "조건 값을 입력하세요.", parent=self)
            return
        if not av:
            messagebox.showwarning("확인", "허용값을 입력하세요.", parent=self)
            return
        self.result = (cc, cv, av, self.msg_e.get().strip())
        self.destroy()


def main():
    root = tk.Tk()
    try:
        ExcelMergeApp(root)
    except Exception:
        import traceback
        traceback.print_exc()
    root.mainloop()


if __name__ == "__main__":
    main()
